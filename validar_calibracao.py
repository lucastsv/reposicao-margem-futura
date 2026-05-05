"""
Validacao do algoritmo de calibracao contra o estado atual da planilha.

Le os valores calculados (H, P, O, B) da aba "Preco Reposicao", roda calibrate(),
e mostra lado a lado: agios atuais vs novos agios, desvios L/O atuais vs novos.

Tolerancia de validacao: linhas-ancora (com cotacao) devem ficar com |L/O - 1| < 1%.
"""

from pathlib import Path
import openpyxl
from calibracao import RowState, calibrate, L, deviation_pct, ANCHOR_ROW, LAST_ROW

XLSX = Path(__file__).parent / "Três Marias - Cálculo de Margem Futura.xlsx"
SHEET = "Preço Reposição"


def read_state() -> list[RowState]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in range(ANCHOR_ROW, LAST_ROW + 1):
        rows.append(RowState(
            row=r,
            B=ws.cell(row=r, column=2).value,   # B
            H=ws.cell(row=r, column=8).value,   # H
            P=ws.cell(row=r, column=16).value,  # P
            O=ws.cell(row=r, column=15).value or 0,  # O (Average)
        ))
    return rows


def fmt_pct(x):
    return f"{x*100:+.2f}%" if x is not None else "    -  "


def main():
    current = read_state()
    new = calibrate(current)

    print(f"{'lin':>3} {'B':>5} {'O':>9} | {'P_atual':>8} {'L_atual':>9} {'desvio':>8} | {'P_novo':>8} {'L_novo':>9} {'desvio':>8}")
    print("-" * 95)
    max_dev_after = 0.0
    anchors_after = 0
    for c, n in zip(current, new):
        anchor_marker = " *" if c.O > 0 else "  "
        dev_c = deviation_pct(c)
        dev_n = deviation_pct(n)
        if dev_n is not None:
            max_dev_after = max(max_dev_after, abs(dev_n))
            anchors_after += 1
        print(
            f"{c.row:>3}{anchor_marker} {c.B:>4.0f} {c.O:>9.2f} | "
            f"{c.P*100:>7.2f}% {L(c):>9.2f} {fmt_pct(dev_c):>8} | "
            f"{n.P*100:>7.2f}% {L(n):>9.2f} {fmt_pct(dev_n):>8}"
        )

    print()
    print(f"Linhas-ancora (* marcadas): {anchors_after}")
    print(f"Maior desvio L/O apos calibracao: {max_dev_after*100:.4f}%")
    print(f"Tolerancia: 1.00%  ->  {'OK' if max_dev_after < 0.01 else 'FALHA'}")


if __name__ == "__main__":
    main()
