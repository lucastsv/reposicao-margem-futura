"""
Pipeline diario: coleta CEPEA, escreve no R da linha mapeada por peso,
recomputa O em Python, recalibra agios via calibrate(), salva o .xlsm.

Uso direto: python atualizar_planilha.py
Pre-requisito: Excel deve estar fechado (o .xlsm fica em uso exclusivo).
"""

from __future__ import annotations
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from cepea_collector import collect as collect_cepea, BezerroSnapshot
from datagro_collector import collect as collect_datagro, DatagroSnapshot
from scot_collector import collect as collect_scot, ScotSnapshot
from calibracao import RowState, calibrate, ANCHOR_ROW, LAST_ROW
from pipeline_core import (
    coleta_com_retry, computar_O, calcular_B, linha_por_peso,
    SCOT_LINHAS, DATAGRO_LINHAS,
)

XLSM = Path(__file__).parent / "planilha" / "Três Marias - Cálculo de Margem Futura.xlsm"
SHEET = "Preço Reposição"
LOG_DIR = Path(__file__).parent / "logs"

# Colunas (1-indexadas): B=2, H=8, O=15, P=16, Q=17, R=18, S=19, T=20
COL_B, COL_H, COL_O, COL_P, COL_Q, COL_R, COL_S, COL_T = 2, 8, 15, 16, 17, 18, 19, 20


def ler_estado(ws) -> tuple[list[RowState], list[dict]]:
    """Le B (calculado), H, P e cotacoes Q/R/S/T. Retorna (rows com O computado, cotacoes_raw)."""
    rows: list[RowState] = []
    cotacoes: list[dict] = []
    for r in range(ANCHOR_ROW, LAST_ROW + 1):
        B = calcular_B(r)
        H = float(ws.cell(row=r, column=COL_H).value or 0)
        P = float(ws.cell(row=r, column=COL_P).value or 0)

        def num(v):
            return float(v) if isinstance(v, (int, float)) and v else None

        q = num(ws.cell(row=r, column=COL_Q).value)
        rr = num(ws.cell(row=r, column=COL_R).value)
        s = num(ws.cell(row=r, column=COL_S).value)
        t = num(ws.cell(row=r, column=COL_T).value)
        O = computar_O(q, rr, s, t)
        rows.append(RowState(row=r, B=B, H=H, P=P, O=O))
        cotacoes.append({"q": q, "r": rr, "s": s, "t": t})
    return rows, cotacoes


def main() -> int:
    LOG_DIR.mkdir(exist_ok=True)
    log_path = LOG_DIR / f"atualizar_{datetime.now():%Y%m%d_%H%M%S}.log"

    def log(msg: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        print(line)
        with log_path.open("a", encoding="utf-8") as f:
            f.write(line + "\n")

    log(f"Inicio. Arquivo alvo: {XLSM.name}")

    if not XLSM.exists():
        log(f"ERRO: arquivo nao encontrado: {XLSM}")
        return 1

    try:
        snap_cepea: BezerroSnapshot = coleta_com_retry("CEPEA", collect_cepea, log)
    except Exception as e:
        log(f"ERRO coleta CEPEA: {e}")
        return 2

    log(f"CEPEA OK: peso={snap_cepea.peso_medio_kg:.1f}kg ({snap_cepea.data_ultimo_peso}) "
        f"media4d={snap_cepea.media_4d_preco:,.2f} ({snap_cepea.data_ultimo_preco})")

    try:
        snap_dg: DatagroSnapshot = coleta_com_retry("DATAGRO", collect_datagro, log)
    except Exception as e:
        log(f"ERRO coleta DATAGRO: {e}")
        return 5

    log(f"DATAGRO OK: boi MS R$ {snap_dg.boi_gordo_ms_arroba:.2f}/@; reposicao MS R$/kg "
        f"des={snap_dg.reposicao_ms['desmama']:.2f} bez={snap_dg.reposicao_ms['bezerro']:.2f} "
        f"gar={snap_dg.reposicao_ms['garrote']:.2f} mag={snap_dg.reposicao_ms['boi_magro']:.2f} "
        f"({snap_dg.data_ref})")

    try:
        snap_scot: ScotSnapshot = coleta_com_retry("SCOT", collect_scot, log)
    except Exception as e:
        log(f"ERRO coleta SCOT: {e}")
        return 6

    log(f"SCOT OK: MS R$/cab des={snap_scot.desmama_cab:,.2f} bez={snap_scot.bezerro_cab:,.2f} "
        f"gar={snap_scot.garrote_cab:,.2f} mag={snap_scot.boi_magro_cab:,.2f} "
        f"({snap_scot.data_ref})")

    # Le valores ja calculados (cotacoes que sao formula precisam estar resolvidas)
    try:
        wb_data = openpyxl.load_workbook(XLSM, data_only=True, keep_vba=True)
    except PermissionError:
        log("ERRO: arquivo em uso (Excel aberto?). Feche o Excel e tente de novo.")
        return 3

    if SHEET not in wb_data.sheetnames:
        log(f"ERRO: aba '{SHEET}' nao encontrada")
        return 4

    rows, cotacoes = ler_estado(wb_data[SHEET])

    # CEPEA: atualiza R da linha mapeada por peso
    idx_cepea = linha_por_peso(snap_cepea.peso_medio_kg, rows)
    linha_cepea = rows[idx_cepea].row
    log(f"CEPEA -> linha {linha_cepea} "
        f"(B={rows[idx_cepea].B:.0f}kg, {rows[idx_cepea].B/30:.1f}@)")
    cotacoes[idx_cepea]["r"] = snap_cepea.media_4d_preco
    rows[idx_cepea].O = computar_O(**cotacoes[idx_cepea])

    # Coluna R eh propriedade do script CEPEA — limpa R em linhas que nao sao
    # a do CEPEA do dia, pra evitar que valor stale do run anterior vire ancora.
    for i, r_state in enumerate(rows):
        if r_state.row != linha_cepea and cotacoes[i]["r"] is not None:
            cotacoes[i]["r"] = None
            rows[i].O = computar_O(**cotacoes[i])

    # DATAGRO: atualiza H7 (ancora) e S das linhas mapeadas com R$/kg * B
    rows[0].H = snap_dg.boi_gordo_ms_arroba  # rows[0] = linha 7 (ancora)

    # Coluna S eh propriedade do script — limpa qualquer S em linha nao-DATAGRO
    # pra evitar ancoras-fantasma da rotina manual antiga.
    target_dg_rows = set(DATAGRO_LINHAS.values())
    for i, r_state in enumerate(rows):
        if r_state.row != ANCHOR_ROW and r_state.row not in target_dg_rows:
            if cotacoes[i]["s"] is not None:
                cotacoes[i]["s"] = None
                rows[i].O = computar_O(**cotacoes[i])

    # Coluna T (SCOT) — limpa em linhas que nao sao SCOT, depois aplica os 4 valores
    # T7 (=L7) eh self-consistent, deixa intocado.
    target_scot_rows = set(SCOT_LINHAS.values())
    for i, r_state in enumerate(rows):
        if r_state.row != ANCHOR_ROW and r_state.row not in target_scot_rows:
            if cotacoes[i]["t"] is not None:
                cotacoes[i]["t"] = None
                rows[i].O = computar_O(**cotacoes[i])

    scot_map = {
        "desmama":   snap_scot.desmama_cab,
        "bezerro":   snap_scot.bezerro_cab,
        "garrote":   snap_scot.garrote_cab,
        "boi_magro": snap_scot.boi_magro_cab,
    }
    for cat, linha in SCOT_LINHAS.items():
        idx_scot = linha - ANCHOR_ROW
        if idx_scot < 0 or idx_scot >= len(rows):
            log(f"AVISO: linha {linha} fora do range, pulando SCOT {cat}")
            continue
        cotacoes[idx_scot]["t"] = scot_map[cat]
        rows[idx_scot].O = computar_O(**cotacoes[idx_scot])

    for cat, linha in DATAGRO_LINHAS.items():
        rkg = snap_dg.reposicao_ms[cat]
        idx_dg = linha - ANCHOR_ROW
        if idx_dg < 0 or idx_dg >= len(rows):
            log(f"AVISO: linha {linha} fora do range, pulando categoria {cat}")
            continue
        s_total = rkg * rows[idx_dg].B
        cotacoes[idx_dg]["s"] = s_total
        rows[idx_dg].O = computar_O(**cotacoes[idx_dg])

    # Calibra com estado atualizado
    new_rows, rebeldes = calibrate(rows)
    if rebeldes:
        log(f"Ancoras rebeldes descartadas (geraria agio negativo): "
            f"linhas {rebeldes}")

    # Reabre wb com formulas (sem data_only) pra escrever
    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    ws = wb[SHEET]

    # H7 (ancora boi gordo MS DATAGRO)
    ws.cell(row=ANCHOR_ROW, column=COL_H).value = snap_dg.boi_gordo_ms_arroba
    log(f"Escrito H{ANCHOR_ROW} = {snap_dg.boi_gordo_ms_arroba:.2f} (DATAGRO boi gordo MS)")

    # Limpa R das linhas que nao sao a CEPEA do dia (sweep) — peso medio CEPEA
    # muda entre runs e desloca a linha alvo, deixando R stale na linha antiga.
    for r in range(ANCHOR_ROW + 1, LAST_ROW + 1):
        if r != linha_cepea:
            celula = ws.cell(row=r, column=COL_R)
            if celula.value is not None:
                log(f"Limpo R{r} (era stale)")
                celula.value = None

    # CEPEA bezerro
    ws.cell(row=linha_cepea, column=COL_R).value = snap_cepea.media_4d_preco
    log(f"Escrito R{linha_cepea} = {snap_cepea.media_4d_preco:,.2f} (CEPEA bezerro media 4d)")

    # Limpa S das linhas que nao sao DATAGRO (sweep)
    for r in range(ANCHOR_ROW + 1, LAST_ROW + 1):
        if r not in target_dg_rows:
            celula = ws.cell(row=r, column=COL_S)
            if celula.value is not None:
                log(f"Limpo S{r} (era stale)")
                celula.value = None

    # Limpa T das linhas que nao sao SCOT (sweep)
    for r in range(ANCHOR_ROW + 1, LAST_ROW + 1):
        if r not in target_scot_rows:
            celula = ws.cell(row=r, column=COL_T)
            if celula.value is not None:
                log(f"Limpo T{r} (era stale)")
                celula.value = None

    # SCOT MS -> R$/cab numeric
    for cat, linha in SCOT_LINHAS.items():
        ws.cell(row=linha, column=COL_T).value = scot_map[cat]
        log(f"Escrito T{linha} = {scot_map[cat]:,.2f} (SCOT {cat})")

    # DATAGRO reposicao -> formulas =rkg*B<row>
    for cat, linha in DATAGRO_LINHAS.items():
        rkg = snap_dg.reposicao_ms[cat]
        ws.cell(row=linha, column=COL_S).value = f"={rkg}*B{linha}"
        log(f"Escrito S{linha} = ={rkg}*B{linha} (DATAGRO {cat})")

    # P calibrado
    for s in new_rows[1:]:
        ws.cell(row=s.row, column=COL_P).value = s.P
    log(f"Escrito P{ANCHOR_ROW + 1}:P{LAST_ROW} (calibrado)")

    try:
        wb.save(XLSM)
    except PermissionError:
        log("ERRO: arquivo em uso no momento do save (Excel/Explorer aberto?). "
            "Mudancas nao foram persistidas.")
        return 7
    log(f"Salvo: {XLSM.name}")

    # Geracao de imagem local (snapshot do range B4:L30 via Excel COM) — preview offline.
    # NAO envia Telegram aqui: o envio fica exclusivo do pipeline_cloud (GH Actions)
    # pra evitar mensagens duplicadas.
    try:
        from gerar_imagem import gerar

        img_path = LOG_DIR / f"preview_{datetime.now():%Y%m%d}.png"
        gerar(XLSM, SHEET, img_path)
        log(f"Imagem local gerada: {img_path.name}")
    except Exception as e:
        log(f"AVISO: Falha na geracao da imagem local: {e}")

    log("Fim.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
